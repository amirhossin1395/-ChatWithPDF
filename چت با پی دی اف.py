#!/usr/bin/env python3
"""
╔══════════════════════════════════════════╗
║  📄💬 Chat with PDF Pro v1.0          ║
║  PDF | Word | Excel | AI Chat         ║
║  AmirHossein Haji Moradkhani          ║
║  Kerman, Iran                          ║
╚══════════════════════════════════════════╝
"""
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
import PyPDF2
import docx
import openpyxl
import requests
import re
import threading
import os

# ════════════════ GOOGLE COLORS ════════════════
WHITE = "#ffffff"
BLUE = "#1a73e8"
GREY = "#f8f9fa"
BORDER = "#dadce0"
TEXT = "#202124"
SURFACE = "#f1f3f4"

# ════════════════ FILE READER ════════════════
class FileReader:
    @staticmethod
    def read_pdf(path):
        text = ""
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text
    
    @staticmethod
    def read_docx(path):
        doc = docx.Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    
    @staticmethod
    def read_xlsx(path):
        wb = openpyxl.load_workbook(path)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"\n── {sheet} ──\n"
            for row in ws.iter_rows(values_only=True):
                text += " | ".join([str(c) for c in row if c]) + "\n"
        return text
    
    @staticmethod
    def read_txt(path):
        with open(path, 'r', encoding='utf-8') as f:
            return f.read()

# ════════════════ AI ENGINE ════════════════
def ask_ai(question, context=""):
    try:
        url = "https://api.deepseek.com/v1/chat/completions"
        data = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": "بر اساس متن داده شده جواب بده. فارسی و روان حرف بزن. اگه سوال درباره فصل خاصیه، محتوای همون فصل رو بگو."},
                {"role": "user", "content": f"متن سند:\n{context[:4000]}\n\nسوال کاربر: {question}"}
            ],
            "temperature": 0.4, "max_tokens": 600
        }
        r = requests.post(url, headers={"Content-Type": "application/json"}, json=data, timeout=20)
        if r.status_code == 200:
            return r.json()["choices"][0]["message"]["content"]
    except:
        pass
    return "⚠️ خطا در اتصال به سرور. لطفاً دوباره تلاش کنید."

# ════════════════ MAIN APP ════════════════
class ChatWithPDF:
    def __init__(self, root):
        self.root = root
        self.root.title("📄💬 Chat with PDF Pro")
        self.root.geometry("900x700")
        self.root.configure(bg=WHITE)
        
        self.document_text = ""
        self.document_name = ""
        
        self.setup_ui()
    
    def setup_ui(self):
        # ════════════ HEADER ════════════
        header = tk.Frame(self.root, bg=WHITE)
        header.pack(fill=tk.X, padx=30, pady=(20,10))
        
        tk.Label(header, text="📄💬 Chat with PDF", font=("Segoe UI", 26, "bold"),
                bg=WHITE, fg=TEXT).pack(side=tk.LEFT)
        tk.Label(header, text="Pro", font=("Segoe UI", 26, "bold"),
                bg=WHITE, fg=BLUE).pack(side=tk.LEFT)
        
        # ════════════ TOOLBAR ════════════
        toolbar = tk.Frame(self.root, bg=SURFACE, height=50)
        toolbar.pack(fill=tk.X, padx=30, pady=(0,15))
        
        file_types = [
            ("📄 PDF", "*.pdf"),
            ("📝 Word", "*.docx"),
            ("📊 Excel", "*.xlsx"),
            ("📃 Text", "*.txt")
        ]
        
        for icon, ftype in file_types:
            btn = tk.Button(toolbar, text=f"{icon}", font=("Segoe UI", 11),
                           bg=SURFACE, fg=TEXT, relief=tk.FLAT, padx=15, pady=8,
                           command=lambda ft=ftype: self.load_file(ft))
            btn.pack(side=tk.LEFT, padx=5)
        
        self.file_label = tk.Label(toolbar, text="📂 هیچ فایلی انتخاب نشده",
                                    bg=SURFACE, fg=TEXT)
        self.file_label.pack(side=tk.LEFT, padx=20)
        
        # ════════════ CHAT AREA ════════════
        self.chat = scrolledtext.ScrolledText(self.root, font=("Segoe UI", 12),
                                               bg=WHITE, fg=TEXT, relief=tk.FLAT,
                                               borderwidth=1, padx=20, pady=15)
        self.chat.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0,15))
        self.chat.insert(tk.END, "📂 لطفاً یک فایل (PDF، Word، Excel یا متن) انتخاب کنید.\n")
        self.chat.insert(tk.END, "💬 بعدش می‌تونید بپرسید: \"فصل ۳ درباره چیه؟\"\n\n")
        self.chat.configure(state=tk.DISABLED)
        
        # ════════════ INPUT ════════════
        input_frame = tk.Frame(self.root, bg=WHITE)
        input_frame.pack(fill=tk.X, padx=30, pady=(0,20))
        
        self.input_entry = tk.Entry(input_frame, font=("Segoe UI", 13),
                                    bg=SURFACE, fg=TEXT, relief=tk.FLAT,
                                    borderwidth=1)
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0,10))
        self.input_entry.bind("<Return>", lambda e: self.send_message())
        
        tk.Button(input_frame, text="📨 ارسال", command=self.send_message,
                 bg=BLUE, fg=WHITE, font=("Segoe UI", 12, "bold"),
                 relief=tk.FLAT, padx=25, pady=8).pack(side=tk.RIGHT)
    
    def add_message(self, sender, text):
        self.chat.configure(state=tk.NORMAL)
        if sender == "user":
            self.chat.insert(tk.END, f"\n👤 شما:\n{text}\n")
        else:
            self.chat.insert(tk.END, f"\n🤖 Chat with PDF:\n{text}\n")
        self.chat.configure(state=tk.DISABLED)
        self.chat.see(tk.END)
    
    def load_file(self, filetype):
        path = filedialog.askopenfilename(filetypes=[("Documents", filetype)])
        if not path:
            return
        
        ext = os.path.splitext(path)[1].lower()
        self.add_message("bot", f"⏳ در حال خوندن فایل...")
        
        def reader():
            if ext == '.pdf':
                text = FileReader.read_pdf(path)
            elif ext == '.docx':
                text = FileReader.read_docx(path)
            elif ext == '.xlsx':
                text = FileReader.read_xlsx(path)
            else:
                text = FileReader.read_txt(path)
            
            self.document_text = text
            self.document_name = os.path.basename(path)
            
            # Count chapters
            chapters = re.findall(r'(فصل\s+\d+[:\s\-]+[^\n]+)', text)
            
            self.root.after(0, lambda: self.file_label.config(
                text=f"📂 {self.document_name}"))
            self.root.after(0, lambda: self.add_message("bot", 
                f"✅ فایل بارگذاری شد!\n"
                f"📄 نام: {self.document_name}\n"
                f"📊 حجم: {len(text):,} کاراکتر\n"
                f"📖 {len(chapters)} فصل پیدا شد\n\n"
                f"💬 حالا بپرسید:\n"
                f"• \"این سند درباره چیه؟\"\n"
                f"• \"فصل ۳ رو توضیح بده\"\n"
                f"• \"کلمات کلیدی رو بگو\""))
        
        threading.Thread(target=reader).start()
    
    def send_message(self):
        question = self.input_entry.get().strip()
        if not question:
            return
        
        if not self.document_text:
            self.add_message("bot", "⚠️ لطفاً اول یک فایل انتخاب کنید!")
            return
        
        self.add_message("user", question)
        self.input_entry.delete(0, tk.END)
        
        # Check for chapter request
        chapter_match = re.search(r'فصل\s+(\d+)', question)
        if chapter_match:
            num = int(chapter_match.group(1))
            pattern = rf'فصل\s+{num}[:\s\-]+[^\n]+'
            ch_match = re.search(pattern, self.document_text)
            if ch_match:
                start = ch_match.start()
                next_ch = re.search(rf'فصل\s+{num+1}[:\s\-]+', self.document_text)
                end = next_ch.start() if next_ch else min(start + 3000, len(self.document_text))
                context = self.document_text[start:end]
            else:
                context = self.document_text[:3000]
        else:
            context = self.document_text[:4000]
        
        def get_answer():
            answer = ask_ai(question, context)
            self.root.after(0, lambda: self.add_message("bot", answer))
        
        threading.Thread(target=get_answer).start()

# ════════════════ RUN ════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app = ChatWithPDF(root)
    root.mainloop()